#!utf8

"""

        ########        ## ######## ##      ##
        ##     ##       ## ##       ##  ##  ##
        ##     ##       ## ##       ##  ##  ##
        ########        ## ######   ##  ##  ##
        ##     ## ##    ## ##       ##  ##  ##
        ##     ## ##    ## ##       ##  ##  ##
        ########   ######  ########  ###  ###


    - Script reads Excel file with scene title and text to display and creates OBS scenes
    - Script creates different kind of scenes based on the amount of text to display
    - Script will create template scenes at the end of created scenes
    - Script is capable of creation "section scenes"

"""

# import needed libraries
import openpyxl
import os
import json
import sys

# set default variables
medium_text_length = 60
big_text_length = 145


def read_xlsx(base_path, input_file, target_file):
    # reading excel file
    file_path = str(input_file)
    # load workbook based on file_path
    wb = openpyxl.load_workbook(file_path)
    # get active work sheet
    ws = wb.active
    # load the template scene collection data
    with open(base_path+'\\json_templates\\scene_collection_template.json', encoding='utf8') as json_file:
        # get scene collection from json file
        scene_collection = json.load(json_file)
        # split target file path and read name into 'tail'
        _, tail = os.path.split(target_file)
        # set name of scene collection based on target file
        scene_collection['name'] = tail.replace('.json', '')
    # read all rows from excel file
    for idx, _ in enumerate(ws.iter_rows()):
        # iterate through all row in the excel except header row
        if idx > 0:
            # if there is data written
            if ws.cell(idx+1, 1).value:
                # start creation of scenes
                scene_title_raw = ws.cell(idx+1, 1).value.replace(' ', '_')
                scene_title = "[S]__" + scene_title_raw
                # set IDs for scenes and their parts
                scene_text_header_ID = scene_title + "_text_header"
                scene_text_ID = scene_title + "_text"
                scene_text_header = ws.cell(idx + 1, 2).value
                scene_text = ws.cell(idx + 1, 3).value
                #
                if scene_title_raw == 'Kategorie':
                    # check for empty content
                    if scene_text_header is None:
                        scene_text_header = scene_title_raw
                    #
                    scene_data = create_divider_scene(
                        base_path, ">>>>>>>>>>> "+scene_text_header+" <<<<<<<<<<")
                    append_json_data_to_sceneset(scene_collection, scene_data, None,
                                                 None, scene_text_header)
                else:
                    #
                    scene_data, text_header_data, text_data = create_json_data(
                        scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title)
                    # add all created JSON block to the scene collection
                    append_json_data_to_sceneset(scene_collection, scene_data, text_header_data,
                                                 text_data, scene_title)
    # write dummy scenes for use if something comes up
    create_dummy_scenes(scene_text, scene_text_header,
                        base_path, scene_text_header_ID, scene_text_ID, scene_title, scene_collection)
    # we gathered all data and appended it to the template scene collection
    with open(target_file, 'w') as f:
        #
        print('Writing scene collection to: ' + target_file)
        # write final scene collection to target file
        json.dump(scene_collection, f, indent=4)


def create_dummy_scenes(scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title, data):
    #
    scene_data = create_divider_scene(
        base_path, ">>>>>>>>>>> Backup Vorlagen <<<<<<<<<<")
    append_json_data_to_sceneset(data, scene_data, None,
                                 None, "divider")
    # start creation of scenes
    scene_title = "[S]____Vorlage_zwei_Zeilen_1"
    scene_text_header_ID = scene_title + "_text_header"
    scene_text_ID = scene_title + "_text"
    scene_text_header = 'Vorlage'
    scene_text = 'Vorlagen Text'
    scene_data, text_header_data, text_data = create_json_data(
        scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title)
    # add all created JSON block to the scene collection
    append_json_data_to_sceneset(data, scene_data, text_header_data,
                                 text_data, scene_title)
    # start creation of scenes
    scene_title = "[S]____Vorlage_zwei_Zeilen_2"
    scene_text_header_ID = scene_title + "_text_header"
    scene_text_ID = scene_title + "_text"
    scene_text_header = 'Vorlage'
    scene_text = 'Vorlagen Text'
    scene_data, text_header_data, text_data = create_json_data(
        scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title)
    # add all created JSON block to the scene collection
    append_json_data_to_sceneset(data, scene_data, text_header_data,
                                 text_data, scene_title)
    #
    # start creation of scenes
    scene_title = "[S]____Vorlage_eine_Zeile_1"
    scene_text_header_ID = scene_title + "_text_header"
    scene_text_ID = scene_title + "_text"
    scene_text_header = 'Vorlage'
    scene_text = None
    scene_data, text_header_data, text_data = create_json_data(
        scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title)
    # add all created JSON block to the scene collection
    append_json_data_to_sceneset(data, scene_data, text_header_data,
                                 text_data, scene_title)
    # start creation of scenes
    scene_title = "[S]____Vorlage_eine_Zeile_2"
    scene_text_header_ID = scene_title + "_text_header"
    scene_text_ID = scene_title + "_text"
    scene_text_header = 'Vorlage'
    scene_text = None
    scene_data, text_header_data, text_data = create_json_data(
        scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title)
    # add all created JSON block to the scene collection
    append_json_data_to_sceneset(data, scene_data, text_header_data,
                                 text_data, scene_title)
    # start creation of scenes
    scene_title = "[S]____Vorlage_Vollbild_1"
    scene_text_header_ID = scene_title + "_text_header"
    scene_text_ID = scene_title + "_text"
    scene_text_header = 'Vorlagen Text'
    scene_text = 'Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text '
    scene_data, text_header_data, text_data = create_json_data(
        scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title)
    # add all created JSON block to the scene collection
    append_json_data_to_sceneset(data, scene_data, text_header_data,
                                 text_data, scene_title)
    # start creation of scenes
    scene_title = "[S]____Vorlage_Vollbild_2"
    scene_text_header_ID = scene_title + "_text_header"
    scene_text_ID = scene_title + "_text"
    scene_text_header = 'Vorlagen Text'
    scene_text = 'Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text Vorlagen Text '
    scene_data, text_header_data, text_data = create_json_data(
        scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title)
    # add all created JSON block to the scene collection
    append_json_data_to_sceneset(data, scene_data, text_header_data,
                                 text_data, scene_title)


def append_json_data_to_sceneset(data, scene_data, text_header_data, text_data, scene_title):
    # add all created JSON block to the scene collection
    data['sources'].append(scene_data)
    if text_header_data:
        data['sources'].append(text_header_data)
    if text_data:
        data['sources'].append(text_data)
    #
    print('Successfully created scene: ' + scene_title)


def create_json_data(scene_text, scene_text_header, base_path, scene_text_header_ID, scene_text_ID, scene_title):
    # init variables
    font_size_header = 64
    font_size_main = 92
    posY = 1145
    color = None
    text_length = 0
    text_data = None
    font_style = "Standard"
    font_flags = 0
    # get text length
    if scene_text is not None:
        text_length = len(scene_text)
    elif scene_text_header is not None:
        text_length = len(scene_text_header)
    # set variable depending on text length
    if text_length >= big_text_length:
        font_size_header = 72
        font_size_main = 64
        color = 4278190080
    # only text has been given
    if scene_text is None or (scene_text_header is None and scene_text is not None):
        font_size_header = font_size_main
        posY = 1170
    # only second column has been filled
    # write text to first header and set text line to None
    if scene_text_header is None and scene_text is not None:
        scene_text_header = scene_text
        scene_text = None
    # check which font style to use, users request 'Fett' text for longer text scenes
    if text_length >= big_text_length:
        font_style = "Fett"
        font_flags = 1
    # create header json
    text_header_data = create_text_data(
        base_path, scene_text_header_ID, scene_text_header, font_size_header, font_flags, font_style, color)
    # create text json
    if scene_text is not None:
        text_data = create_text_data(
            base_path, scene_text_ID, scene_text, font_size_main, font_flags, font_style, color)
    # create normal scene for short texts
    if text_length < medium_text_length:
        scene_data = create_scene_data(
            base_path, scene_title, scene_text_ID, scene_text_header_ID, posY)
    # medium text length
    elif text_length >= medium_text_length and text_length < big_text_length:
        # DEBUG:
        print(">>>>>> MEDIUM !!!!")
        scene_data = create_scene_data_medium_text(
            base_path, scene_title, scene_text_ID, scene_text_header_ID)
    # scene for long texts
    else:
        scene_data = create_scene_data_big_text(
            base_path, scene_title, scene_text_ID, scene_text_header_ID)
    # return created json types
    return scene_data, text_header_data, text_data


def create_scene_data_medium_text(base_path, scene_title, scene_text_ID, scene_text_header_ID):
    with open(base_path+'\\json_templates\\scene_template.json', encoding='utf8') as json_file:
        data = json.load(json_file)
        data['name'] = scene_title
        data['settings']['items'][4]['name'] = scene_text_ID
        data['settings']['items'][5]['name'] = scene_text_header_ID
        # logo
        data['settings']['items'][2]['pos']['y'] = 1038
        # bg color band
        data['settings']['items'][2]['pos']['y'] = 995
        data['settings']['items'][2]['scale']['y'] = 2.75
        # text
        data['settings']['items'][4]['pos']['y'] = 1097
        # text header
        data['settings']['items'][5]['pos']['y'] = 1038
        return data


def create_scene_data_big_text(base_path, scene_title, scene_text_ID, scene_text_header_ID):
    with open(base_path+'\\json_templates\\scene_big_text_template.json', encoding='utf8') as json_file:
        data = json.load(json_file)
        data['name'] = scene_title
        data['settings']['items'][6]['name'] = scene_text_ID
        data['settings']['items'][7]['name'] = scene_text_header_ID
        # adjust abendmahl scenes
        if any(key.casefold() in str(scene_title).casefold() for key in ["abendmahl", "abendmal"]):
            data['settings']['items'][2]['visible'] = False
            data['settings']['items'][3]['visible'] = False
        return data


def create_divider_scene(base_path, scene_title):
    with open(base_path+'\\json_templates\\divider_template.json', encoding='utf8') as json_file:
        data = json.load(json_file)
        data['name'] = scene_title
        return data


def create_scene_data(base_path, scene_title, scene_text_ID, scene_text_header_ID, posY):
    with open(base_path+'\\json_templates\\scene_template.json', encoding='utf8') as json_file:
        data = json.load(json_file)
        data['name'] = scene_title
        data['settings']['items'][4]['name'] = scene_text_ID
        data['settings']['items'][5]['name'] = scene_text_header_ID
        data['settings']['items'][5]['pos']['y'] = posY
        return data


def create_text_data(base_path, scene_text_ID, scene_text, font_size, font_flags, font_style, color):
    with open(base_path+'\\json_templates\\text_template.json', encoding='utf8') as json_file:
        data = json.load(json_file)
        data['name'] = scene_text_ID
        data['settings']['text'] = scene_text
        data['settings']['font']['size'] = font_size
        data['settings']['font']['flags'] = font_flags
        data['settings']['font']['style'] = font_style
        if color:
            data['settings']['color'] = color
        return data


if __name__ == "__main__":
    #
    print('Welcome to the automated scene creation by Bjew')
    #
    try:
        if sys.argv and len(sys.argv) > 2:
            #
            input_file = sys.argv[1]
            target_file = sys.argv[2]
            #
            base_path = os.path.dirname(os.path.abspath(sys.argv[0]))
            read_xlsx(base_path, input_file, target_file)
            #
            print('Success!')
        else:
            #
            print('Missing parameters!')
    except Exception as e:
        print(e)
